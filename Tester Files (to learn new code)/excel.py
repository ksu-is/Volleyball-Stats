from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


window = Tk()
window.geometry('1000x1000')
window.title("Volleyball Statistics")
window.configure(background = "white")

# Create Workbook Instance
wb = Workbook()
# Load Existing Workbook
wb = load_workbook("Player Stats 2021 - 2022 (13 Elite)1.xlsx")

# Create active worksheet
ws = wb.active

# Create Variable for Column A
column_a = ws['A']
column_b = ws['B']

print(column_a)

def get_a():
    list = ''
    for cell in column_a:
        list = f'{list + str(cell.value)}\n'
    label_a.config(text = list)
def get_b():
    list = ''
    for cell in column_b:
        list = f'{list + str(cell.value)}\n'
    label_b.config(text = list)

# Column A Button & Label
columna_btn = Button(window, text = "Get Column A", width = 12, command = get_a)
columna_btn.grid(row = 0, column = 0)

label_a = Label(window, text = "")
label_a.grid(row = 1, column = 0)
# Column B Button & Label
columnb_btn = Button(window, text = "Get Column B", width = 12, command = get_b)
columnb_btn.grid(row = 0, column = 1)

label_b = Label(window, text = "")
label_b.grid(row = 1, column = 1)

# Add Raven Watts # 2 to A14 and B14
ws['A14'] = "Raven Watts"
ws['B14'] = "2"

# Save File
wb.save('Player Stats 2021 - 2022 (13 Elite)1.xlsx')



window.mainloop()
