from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook("Tester.xlsx")

ws = wb.active

# view cell content
#print(ws['B2'].value)


# change values in excel
#ws['A2'] = "Aniyah"

# Need to save the workbook to create changes
# make sure excel sheet is closed 
#wb.save("Tester.xlsx")

# View all of the sheet names
#print(wb.sheetnames)

#print(wb['Sheet1'])

#create a sheet 
#wb.create_sheet("Test")
#print(wb.sheetnames)
#wb.save("Tester.xlsx")



#ws.title = "Data"
#ws.append(['Tim', 'is','great'])
#wb.save("Tester.xlsx")



# loop through rows and columns
#from openpyxl.utils import get_column_letter


#for row in range(1, 11):
 #   for col in range(1, 5):
    # for col in range(0, 4):    
        # gives character repersented by 65 - Capital A
        #char = chr(65 + col)
  #      char = get_column_letter(col)
   #     print(ws[char + str(row)].value)
