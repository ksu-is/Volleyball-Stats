# My Vollyeyball project

from tkinter import *
#from tkinter import ttk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Creating Framework
window = Tk()
window.geometry('1000x1000')
window.title("Volleyball Statistics")
window.configure(background = "white")

wb = Workbook()
filename = 'Volleyball Stats.xlsx'
wb = load_workbook(filename)
ws = wb.active


#stats = ttk.Treeview(window)

# Creating a Logo 
logo = PhotoImage(file="Volleyball Stats.gif")
Label (window, image=logo, bg = "white").grid(row=0, columnspan = 3)


def profile():
 # CREATING "LOGIN" FUNCTION
   def login():
   # formatting 
      log_in = Toplevel()
      log_in.geometry("500x500")
      log_in.title("Login Page")
   # Creating Labels
      email_phone = Label(log_in, text = "Enter your email or phone number: ")
      password = Label(log_in, text = "Enter your password: ")
   # Packing Labels
      email_phone.grid(row = 2, column = 0)
      password.grid(row = 3, column = 0)
   # Variable for storing data
      email_phonevalue = StringVar
      passwordvalue = StringVar
   # Creating Entry Fields
      email_phoneentry = Entry(log_in, textvariable = email_phonevalue)
      passwordentry = Entry(log_in, textvariable = passwordvalue)
   # Packing Entry Fields
      email_phoneentry.grid(row = 2, column = 1)
      passwordentry.grid(row = 3, column = 1)
      passwordentry.config(show = "*")
   # defining enter button
      def enter():
         username = email_phoneentry.get()
         password = passwordentry.get()
         print("Username," , username, "and Password, ", password, ", successful" )


         email_phoneentry.delete(0, END)
         passwordentry.delete(0, END)

         confirmation = Label(log_in, text = "Log In was Successful")
         confirmation.grid(row = 6, columnspan= 3)
   # Enter Button
      Button(log_in, text = "Enter", command = enter).grid(row = 4, column = 1)
   # CREATING A "CREATE A PROFILE" FUNCTION
      def createp():
         create_profile = Toplevel()
         create_profile.geometry("500x500")
         create_profile.title("Create a Profile")

         name = Label(create_profile, text = "Enter your name: ")    
         email = Label(create_profile, text = "Enter your email: ")
         phone = Label(create_profile, text = "Enter your phone number: ")
         password = Label(create_profile, text = "Enter your password: ")
         confirm_password = Label(create_profile, text = "Confirm your password: ")
      # Packing Labels
         name.grid(row = 1, column = 0)
         email.grid(row = 2, column = 0)
         phone.grid(row = 3, column = 0)
         password.grid(row = 4, column = 0)
         confirm_password.grid(row = 5, column = 0)
      # Variable for storing data
         namevalue = StringVar
         emailvalue = StringVar
         phonevalue = StringVar
         passwordvalue = StringVar
         confirm_passwordvalue = StringVar
         checkvalue = IntVar
         checkvalue2 = IntVar
      # Creating Entry Fields
         nameentry = Entry(create_profile, textvariable = namevalue)
         emailentry = Entry(create_profile, textvariable = emailvalue)
         phoneentry = Entry(create_profile, textvariable = phonevalue)
         passwordentry = Entry(create_profile, textvariable = passwordvalue)
         confirm_passwordentry = Entry(create_profile, textvariable = confirm_passwordvalue)
      # Packing Entry Fields
         nameentry.grid(row = 1, column = 1)
         emailentry.grid(row = 2, column = 1)
         phoneentry.grid(row = 3, column = 1)
         passwordentry.grid(row = 4, column = 1)
         confirm_passwordentry.grid(row = 5, column = 1) 
      # Creating Checkbox
         checkbtn = Checkbutton(create_profile, text = "Coach", variable = checkvalue)
         checkbtn.grid(row=6, column = 0)

         checkbtn2 = Checkbutton(create_profile, text = "Player", variable = checkvalue2)
         checkbtn2.grid(row=6, column = 1)
      # defining submit button
         def profile():
               name = nameentry.get()
               email = emailentry.get()
               phone = phoneentry.get()
               password = passwordentry.get()
               confirm = confirm_passwordentry.get()

               nameentry.delete(0, END)
               emailentry.delete(0, END)
               phoneentry.delete(0, END)
               passwordentry.delete(0, END)
               confirm_passwordentry.delete(0, END)

               print("Team creation successful. Information below:\n",name, email,phone,password,confirm,coachcheck,playercheck)
               confirmation = Label(create_profile, text = "New profile was created! Please proceed to login page" )
               confirmation.grid(row = 8 , column = 1)
               
               
      # Submit Button
         Button(create_profile, text = "Submit", command = profile).grid(row = 7, column = 1)
   # Creating a "Create a profile" button
      create_bttn = Button (log_in, text = "Create a New Profile", width = 15 , command = createp)
      create_bttn.grid(row = 4 , column = 0)


   # Creating a "Login" Button
   login_bttn = Button (window, text = "Login", width = 15 , command = login)
   login_bttn.grid(row = 2 , column = 0)
profile()

def teams():
 # CREATING A "MY TEAMS" FUNCTION
   def my_teams():
   # formatting 
      myteams = Toplevel()
      myteams.geometry("500x500")
      myteams.title("My Teams")
   # Creating Header / Explanation
      myteam_lbl = Label(myteams, text = "My Teams!", font = "arial 12 bold", bg = "white" )
      myteam_lbl.grid(row = 0 , columnspan= 3)
      def create():
       # CREATING A "CREATE TEAM" FUNCTION
         def create_team():
         # formatting 
            new_team = Toplevel()
            new_team.geometry("500x500")
            new_team.title("Create a New Team")
         # Creating Header / Explanation
            new_team_lbl = Label(new_team, text = "Create a New Team!", font = "arial 12 bold", bg = "white" )
            new_team_lbl.grid(row = 1 , columnspan= 3)

         # Enter Team Name
            team_name_lbl = Label(new_team, text = "Enter Team Name:", font = "arial 10", bg = "white" )
            team_name_lbl.grid(row = 2 , column = 0)

            team_name = Entry(new_team, width = 30, bg = "white")
            team_name.grid(row = 2, column = 1)
         # Creating Jersey and Player Number Fields
            jersey_lbl = Label(new_team, text = "Enter player's Jersey Number: ", font = "arial 9", bg = "white")
            player_lbl = Label(new_team, text = "Enter player's Name: ", font = "arial 9", bg = "white")
         # Packing Fields
            jersey_lbl.grid(row = 3 , column =0)
            player_lbl.grid(row = 3, column =1)
         # Creating Player Name and Jersey Entry Boxes
            player1 = Entry(new_team, width = 30, bg = "white")
            player2 = Entry(new_team, width = 30, bg = "white")
            player3 = Entry(new_team, width = 30, bg ="white")
            player4 = Entry(new_team, width = 30, bg = "white")
            player5 = Entry(new_team, width = 30, bg = "white")
            player6 = Entry(new_team, width = 30, bg = "white")
            player7 = Entry(new_team, width = 30, bg = "white")
            player8 = Entry(new_team, width = 30, bg = "white")
            player9 = Entry(new_team, width = 30, bg = "white")
            player10 = Entry(new_team, width = 30, bg = "white")
            player11 = Entry(new_team, width = 30, bg = "white")
            player12 = Entry(new_team, width = 30, bg = "white")
            player13 = Entry(new_team, width = 30, bg = "white")
            player14 = Entry(new_team, width = 30, bg = "white")
            player15 = Entry(new_team, width = 30, bg = "white")
            player16 = Entry(new_team, width = 30, bg = "white")
         #Jerseys Entry
            jersey1 = Entry(new_team, width = 10, bg = "white")
            jersey2 = Entry(new_team, width = 10, bg = "white")
            jersey3 = Entry(new_team, width = 10, bg = "white")
            jersey4 = Entry(new_team, width = 10, bg = "white")
            jersey5 = Entry(new_team, width = 10, bg = "white")
            jersey6 = Entry(new_team, width = 10, bg = "white")
            jersey7 = Entry(new_team, width = 10, bg = "white")
            jersey8 = Entry(new_team, width = 10, bg ="white")
            jersey9 = Entry(new_team, width = 10, bg = "white")
            jersey10 = Entry(new_team, width = 10, bg = "white")
            jersey11 = Entry(new_team, width = 10, bg = "white")
            jersey12 = Entry(new_team, width = 10, bg = "white")
            jersey13 = Entry(new_team, width = 10, bg = "white")
            jersey14 = Entry(new_team, width = 10, bg = "white")
            jersey15 = Entry(new_team, width = 10, bg = "white")
            jersey16 = Entry(new_team, width = 10, bg = "white")
         # Packing Jersey and Player Fields 
          # Player Packing
            player1.grid(row = 4, column = 1)
            player2.grid(row = 5, column = 1)
            player3.grid(row = 6, column = 1)
            player4.grid(row = 7, column = 1)
            player5.grid(row = 8, column = 1)
            player6.grid(row = 9, column = 1)
            player7.grid(row = 10, column = 1)
            player8.grid(row = 11, column = 1)
            player9.grid(row = 12, column = 1)
            player10.grid(row = 13, column = 1)
            player11.grid(row = 14, column = 1)
            player12.grid(row = 15, column = 1)
            player13.grid(row = 16, column = 1)
            player14.grid(row =17, column = 1)
            player15.grid(row = 18, column = 1)
            player16.grid(row = 19, column = 1)
          # Jersey Packing
            jersey1.grid(row = 4, column = 0)
            jersey2.grid(row = 5, column = 0)
            jersey3.grid(row = 6, column = 0)
            jersey4.grid(row = 7, column = 0)
            jersey5.grid(row = 8, column = 0)
            jersey6.grid(row =9, column = 0)
            jersey7.grid(row = 10, column = 0)
            jersey8.grid(row = 11, column = 0)
            jersey9.grid(row = 12, column = 0)
            jersey10.grid(row = 13, column = 0)
            jersey11.grid(row = 14, column = 0)
            jersey12.grid(row = 15, column = 0)
            jersey13.grid(row = 16, column = 0)
            jersey14.grid(row = 17, column = 0)
            jersey15.grid(row = 18, column = 0)
            jersey16.grid(row = 19, column = 0)
         # Submit & Cancel Button Functions
            def submit():
               players = []
               numbers = []
            #Retrieve and print player names & numbers
               name1 = player1.get()
               num1 = jersey1.get()
               name2 = player2.get()
               num2 = jersey2.get()
               name3 = player3.get()
               num3 = jersey3.get()
               name4 = player4.get()
               num4 = jersey4.get()
               name5 = player1.get()
               num5 = jersey1.get()
               name6 = player6.get()
               num6 = jersey6.get()
               name7 = player7.get()
               num7 = jersey7.get()
               name8 = player8.get()
               num8 = jersey8.get()
               name9 = player9.get()
               num9 = jersey9.get()
               name10 = player10.get()
               num10 = jersey10.get()
               name11 = player11.get()
               num11 = jersey11.get()
               name12 = player12.get()
               num12 = jersey12.get()
               name13 = player13.get()
               num13 = jersey13.get()
               name14 = player14.get()
               num14 = jersey14.get()
               name15 = player15.get()
               num15 = jersey15.get()
               name16 = player16.get()
               num16 = jersey16.get()
            
            # Creating a window to store the data in a list
             # players
               players.append(name1)
               players.append(name2)
               players.append(name3)
               players.append(name4)
               players.append(name5)
               players.append(name6)
               players.append(name7)
               players.append(name8)
               players.append(name9)
               players.append(name10)
               players.append(name11)
               players.append(name12)
               players.append(name13)
               players.append(name14)
               players.append(name15)
               players.append(name16)
             # Numbers
               numbers.append(num1)
               numbers.append(num2)
               numbers.append(num3)
               numbers.append(num4)
               numbers.append(num5)
               numbers.append(num6)
               numbers.append(num7)
               numbers.append(num8)
               numbers.append(num9)
               numbers.append(num10)
               numbers.append(num11)
               numbers.append(num12)
               numbers.append(num13)
               numbers.append(num14)
               numbers.append(num15)
               numbers.append(num16)
            
            # Define Team Data Function
               def teamdata():
                  team_data = Toplevel()
                  team_data.geometry("500x500")
                  team_data.title(get_teamname)
                  # pull data from excel sheet 

               get_teamname = team_name.get()
               teamname = Button(myteams,text = get_teamname, width = 30, command = teamdata)
               teamname.grid()

             # Creating a new sheet for data in excel  
               wb.create_sheet(get_teamname)
               print(wb.sheetnames)
               wb.save(filename)

               # turn excel sheet into table in another window inside of each team
            #Clear Entries
               
               team_name.delete(0,END)

               player1.delete(0,END)
               player2.delete(0,END)
               player3.delete(0,END)
               player4.delete(0,END)
               player5.delete(0,END)
               player6.delete(0,END)
               player7.delete(0,END)
               player8.delete(0,END)
               player9.delete(0,END)
               player10.delete(0,END)
               player11.delete(0,END)
               player12.delete(0,END)
               player13.delete(0,END)
               player14.delete(0,END)
               player15.delete(0,END)
               player16.delete(0,END)

               jersey1.delete(0,END)
               jersey2.delete(0,END)
               jersey3.delete(0,END)
               jersey4.delete(0,END)
               jersey5.delete(0,END)
               jersey6.delete(0,END)
               jersey7.delete(0,END)
               jersey8.delete(0,END)
               jersey9.delete(0,END)
               jersey10.delete(0,END)
               jersey11.delete(0,END)
               jersey12.delete(0,END)
               jersey13.delete(0,END)
               jersey14.delete(0,END)
               jersey15.delete(0,END)
               jersey16.delete(0,END)

            # Creating a confirmation statement
               confirmation = Label(new_team, text = "New Team Creation Successful! Please Exit or Create Another Team", font = "arial 9", bg = 'white')
               confirmation.grid(row = 21, columnspan = 3)

               exit = Button(new_team,text = "Exit", width = 30, command = new_team.destroy)
               exit.grid(row = 22, column = 0)
               
         # Creating Submit and Cancel Buttons
            submit_btn = Button(new_team, text = "Submit", command = submit)
            submit_btn.grid(row = 20 , column = 1)
            cancel_btn = Button(new_team, text = "Cancel", command = new_team.destroy)
            cancel_btn.grid(row = 20 , column = 0)
         # Creating a "Create a Team" Button
         newteam_bttn = Button (myteams, text = "Create a New Team", width = 15 , command = create_team)
         newteam_bttn.grid(column = 0)
      create()
   # Creating a "My Teams" Button
   myteams_bttn = Button (window, text = "My Teams", width = 15 , command = my_teams)
   myteams_bttn.grid(row = 2 , column = 1)
teams()


window.mainloop()


